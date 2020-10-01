import csv
from openpyxl import load_workbook
from collections import defaultdict
import logging
from typing import Union, Iterable

logger = logging.getLogger("adjust-scan-images")


SETTING_KEYS_DEFAULT = {
    "resize_ratio": 1,
    "is_align": 1,
    "marker_range": 200,
    "marker_gaussian_ksize": 15,
    "marker_gaussian_std": 3,
    "is_marksheet": 1,
    "is_marksheet_fit": 1,
    "sheet_coord_style": "circle",
    "sheet_gaussian_ksize": 15,
    "sheet_gaussian_std": 3,
}

def read_metadata(
    filepath: Union[None, str] = None,
    mode: str = "excel",
    excel_sheet_name: str = "image_setting",
    **kargs,
) -> dict:
    """
    Metadata setting loader.

    Parameters
    ----------
    filepath : str
        Setting file path.
    mode : str, optional
        Mode, by default "excel"
    excel_sheet_name : str, optional
        Excel sheet name, by default "image_setting"

    Returns
    -------
    dict
        Metadata.
    """
    pre_metadata = {}
    if filepath:
        wb = load_workbook(filepath, read_only=True)
        ws = wb[excel_sheet_name]
        for row in ws.iter_rows(min_row=3):
            key, value = row[0].value, row[1].value
            pre_metadata[key] = value

    # put default value
    for key, value in SETTING_KEYS_DEFAULT.items():
        if key not in pre_metadata:
            pre_metadata[key] = value

    # logging
    logger.info(f"Metadata loaded: {pre_metadata}")

    # formatting
    metadata = {}
    scale = metadata["resize_ratio"] = float(pre_metadata["resize_ratio"])
    metadata["is_marksheet"] = int(pre_metadata["is_marksheet"])
    v = int(float((pre_metadata["marker_range"])) * scale)
    metadata["marker_range"] = (
        (0, 0, v, v),
        (0, -v, v, v),
        (-v, -v, v, v),
        (-v, 0, v, v),
    )
    metadata["is_align"] = int(pre_metadata["is_align"])
    metadata["marker_gaussian_ksize"] = int(
        int(pre_metadata["marker_gaussian_ksize"]) * scale
    )
    metadata["marker_gaussian_std"] = int(
        int(pre_metadata["marker_gaussian_std"]) * scale
    )
    metadata["is_marksheet"] = int(pre_metadata["is_marksheet"])
    metadata["is_marksheet_fit"] = int(pre_metadata["is_marksheet_fit"])
    metadata["sheet_coord_style"] = pre_metadata["sheet_coord_style"]
    metadata["sheet_gaussian_ksize"] = int(
        int(pre_metadata["sheet_gaussian_ksize"]) * scale
    )
    metadata["sheet_gaussian_std"] = int(
        int(pre_metadata["sheet_gaussian_std"]) * scale
    )
    logger.debug(f"Metadata formatted: {metadata}")

    return metadata


def read_marksheet_setting(
    filepath: str,
    scale: float = 1,
    mode: str = "excel",
    excel_sheet_name: str = "marksheet",
    **kargs,
) -> dict:
    """
    Read Marksheet Setting.

    Parameters
    ----------
    filepath : str
        File path.
    scale : float, optional
        Rescale ratio. This corresponds to resize ratio, by default None
    mode : str, optional
        Mode, by default "excel"
    excel_sheet_name : str, optional
        Excel sheet name, by default "marksheet"

    Returns
    -------
    dict
        Marksheet data.
    """
    wb = load_workbook(filepath, read_only=True)
    ws = wb[excel_sheet_name]
    marks = defaultdict(dict)
    for row in ws.iter_rows(min_row=3):
        category, value, x, y, r = [v.value for v in row]
        x, y, r = int(float(x * scale)), int(float(y * scale)), int(float(r * scale))
        marks[category][value] = (x, y, r)
    marks = dict(marks)
    logger.info(f"marksheet data loaded: {marks}")
    return marks


class MarksheetResultWriter:
    def __init__(self, filepath: str, header: Iterable[str]):
        self.f = open(filepath, "a", encoding="shift_jis", newline="")
        self.is_open = True
        self.writer = csv.DictWriter(self.f, header, extrasaction="ignore")
        self.writer.writeheader()

    def write_one_dict(self, data: dict):
        self.writer.writerow(data)

    def close(self):
        self.f.close()
        self.is_open = False
