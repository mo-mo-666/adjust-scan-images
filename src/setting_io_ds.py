from openpyxl import load_workbook
import os
from collections import defaultdict
import logging
from typing import Union, Iterable

logger = logging.getLogger("adjust-scan-images")

SETTING_KEYS_DEFAULT = {
    "resize_ratio": 1,
    "coord_unit": "pt",
    "is_align": 1,
    "marker_range": 200 / 300 * 72,
    "marker_gaussian_ksize": 15,
    "marker_gaussian_std": 3,
    "is_marksheet": 1,
    "is_marksheet_fit": 1,
    "sheet_coord_style": "bbox",
    "sheet_score_threshold": 0,
    "sheet_gaussian_ksize": 15,
    "sheet_gaussian_std": 3,
}

MARK_CATEGORIES = ("room", "class", "student_number_10", "student_number_1")


def read_metadata(
    filepath: Union[None, str] = None,
    mode: str = "excel",
    excel_sheet_name: str = "image_setting",
    pt2px: Union[None, float] = None,
    *args,
    **kargs,
) -> dict:
    """
    Metadata setting loader.

    Parameters
    ----------
    filepath : str
        Setting file path.
    mode : str, optional
        Mode, by default "excel"
    excel_sheet_name : str, optional
        Excel sheet name, by default "image_setting"

    Returns
    -------
    dict
        Metadata.
    """
    metadata = {}
    if filepath:
        wb = load_workbook(filepath, read_only=True)
        ws = wb[excel_sheet_name]
        for row in ws.iter_rows(min_row=3):
            key, value = row[0].value, row[1].value
            metadata[key] = value

    # put default value
    for key, value in SETTING_KEYS_DEFAULT.items():
        if key not in metadata:
            metadata[key] = value

    # logging
    logger.info(f"Metadata loaded: {metadata}")

    # formatting
    metadata["resize_ratio"] = scale = float(metadata["resize_ratio"])
    metadata["is_marksheet"] = int(metadata["is_marksheet"])
    if metadata.get("coord_unit", "px") == "pt" and pt2px is not None:
        v = int(float((metadata["marker_range"])) * scale * (pt2px * scale) / 72)
    else:
        v = int(float((metadata["marker_range"])) * scale)
    metadata["marker_range"] = (
        (0, 0, v, v),
        (0, -v, v, v),
        (-v, -v, v, v),
        (-v, 0, v, v),
    )
    metadata["is_align"] = int(metadata["is_align"])
    metadata["marker_gaussian_ksize"] = int(
        int(metadata["marker_gaussian_ksize"]) * scale
    )
    metadata["marker_gaussian_std"] = int(int(metadata["marker_gaussian_std"]) * scale)
    metadata["is_marksheet"] = int(metadata["is_marksheet"])
    metadata["is_marksheet_fit"] = int(metadata["is_marksheet_fit"])
    metadata["sheet_score_threshold"] = float(metadata["sheet_score_threshold"])
    metadata["sheet_gaussian_ksize"] = int(
        int(metadata["sheet_gaussian_ksize"]) * scale
    )
    metadata["sheet_gaussian_std"] = int(int(metadata["sheet_gaussian_std"]) * scale)
    logger.debug(f"Metadata formatted: {metadata}")

    return metadata


def read_marksheet_setting(
    filepath: Union[None, str],
    scale: float = 1,
    categories: Union[None, Iterable[str]] = MARK_CATEGORIES,
    pt2px: Union[None, float] = None,
    *args,
    **kargs,
) -> dict:
    if not filepath or not categories:
        return {}
    wb = load_workbook(filepath, read_only=True, data_only=True)

    if not pt2px:
        pt2px = 72
    marks = defaultdict(dict)
    for category in categories:
        cell_range = wb.defined_names[category].destinations
        if not cell_range:
            continue
        cells = [wb[s][r] for s, r in cell_range][0]
        for row in cells:
            data = [r.value for r in row]
            assert (
                len(data) == 5
            ), f"The excel data length of {data} is {len(data)} != 5. The format must be (value, x1, y1, x2, y2) for each row."
            value, x1, y1, x2, y2 = data
            x1, y1, x2, y2 = [
                int(float(z) * scale * (pt2px * scale) / 72) for z in (x1, y1, x2, y2)
            ]
            marks[category][value] = (x1, y1, x2, y2)
    marks = dict(marks)
    logger.info(f"marksheet data loaded: {marks}")
    return marks


def decide_save_filepath(
    read_path: str, save_dir: str, data: Union[dict, None] = None
) -> str:
    """
    Decide file name when saving an image. OVERRIDE THIS TO CHANGE THE FILENAME.

    Parameters
    ----------
    read_path : str
        Original file path.
    save_dir: str
        Save directory name.
    data : Union[dict, None], optional
        The data used by deciding the file name, by default None.

    Returns
    -------
    str
        Save file path.
    """
    read_filename = os.path.basename(read_path)
    read_dir = os.path.basename(os.path.dirname(read_path))
    if data:
        _, ext = os.path.splitext(read_filename)
        save_filename = f"{read_dir}-{data.get('room', 'x')}_{data.get('class', 'x')}_{data.get('student_number_10', 'x')}{data.get('student_number_1','x')}{ext}"
    else:
        save_filename = read_filename
    return save_filename
